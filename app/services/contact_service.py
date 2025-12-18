import json
import os
import aiofiles
import logging
from typing import List, Dict
from datetime import datetime
from app.core.config import settings
from app.models.contact import Contact
from fastapi import HTTPException
from app.utils.contact_helpers import normalize_phone, clean_name

logger = logging.getLogger(__name__)

class ContactService:
    def __init__(self):
        self.file_path = os.path.join(settings.DATA_DIR, "contacts.json")
        self._ensure_file_exists()
        # Contact service no longer uses LLM directly - it uses llm_service instead
        self.configured = True
        # Кэш для быстрого поиска по telegram_user_id
        self._telegram_id_cache: Dict[int, Contact] = {}
        self._cache_loaded = False

    def _ensure_file_exists(self):
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        if not os.path.exists(self.file_path):
            with open(self.file_path, 'w', encoding='utf-8') as f:
                json.dump([], f)

    async def get_contacts(self) -> List[Contact]:
        print(f"[ContactService] Reading from: {self.file_path}")
        print(f"[ContactService] File exists: {os.path.exists(self.file_path)}")
        
        async with aiofiles.open(self.file_path, 'r', encoding='utf-8') as f:
            content = await f.read()
            print(f"[ContactService] File content length: {len(content)}")
            
            # Handle empty file
            if not content or not content.strip():
                print("[ContactService] File is empty")
                return []
            try:
                data = json.loads(content)
                print(f"[ContactService] Loaded {len(data)} contacts from JSON")
                contacts = []
                has_changes = False
                seen_ids = set()  # Для дедупликации по ID
                duplicates_removed = 0
                
                for item in data:
                    # Дедупликация по ID - пропускаем дубликаты
                    contact_id = item.get("id")
                    if contact_id is not None and contact_id in seen_ids:
                        duplicates_removed += 1
                        has_changes = True
                        continue
                    if contact_id is not None:
                        seen_ids.add(contact_id)
                    
                    # Migration: If group is missing, set to "Тренеры по футболу"
                    if "group" not in item:
                        item["group"] = "Тренеры по футболу"
                        has_changes = True
                    # Migration: Initialize analyzed_message_ids if missing
                    if "analyzed_message_ids" not in item:
                        item["analyzed_message_ids"] = []
                        has_changes = True
                    contacts.append(Contact(**item))
                
                if duplicates_removed > 0:
                    logger.info(f"🧹 Removed {duplicates_removed} duplicate contacts by ID")
                    print(f"[ContactService] Removed {duplicates_removed} duplicate contacts")
                
                # Save back if migration happened or duplicates removed
                if has_changes:
                    print(f"[ContactService] Saving {len(contacts)} contacts after cleanup")
                    await self._save_data(contacts)
                    
                print(f"[ContactService] Returning {len(contacts)} contacts")
                return contacts
            except json.JSONDecodeError as e:
                print(f"[ContactService] JSON decode error: {e}")
                print(f"[ContactService] Content preview: {content[:500]}")
                return []

    async def get_contact_by_id(self, contact_id: int) -> Contact:
        """Get a single contact by ID"""
        contacts = await self.get_contacts()
        for contact in contacts:
            if contact.id == contact_id:
                return contact
        return None
    
    async def get_contact_by_telegram_id(self, telegram_user_id: int) -> Contact:
        """Быстрый поиск контакта по Telegram user ID"""
        contacts = await self.get_contacts()
        
        # Ищем по полю telegram_user_id
        for contact in contacts:
            if contact.telegram_user_id == telegram_user_id:
                return contact
        
        # Fallback: ищем в extracted_info
        telegram_id_str = f"Telegram ID: {telegram_user_id}"
        for contact in contacts:
            if contact.extracted_info and telegram_id_str in contact.extracted_info:
                # Обновляем поле для будущих запросов
                contact.telegram_user_id = telegram_user_id
                await self.update_contact(contact.id, contact)
                return contact
        
        return None

    async def save_contacts(self, new_contacts: List[Contact]):
        existing_contacts = await self.get_contacts()
        
        # Determine next ID
        next_id = max([c.id for c in existing_contacts if c.id is not None], default=0) + 1
        
        # Track phone numbers to prevent duplicates (используем нормализованные номера)
        seen_phones = {normalize_phone(c.phone) for c in existing_contacts if c.phone}
        
        for contact in new_contacts:
            if contact.id is None:
                contact.id = next_id
                next_id += 1
            
            # Нормализуем телефон перед сохранением
            if contact.phone:
                normalized_phone = normalize_phone(contact.phone)
                if normalized_phone:
                    contact.phone = normalized_phone
                else:
                    # Если нормализация не удалась, оставляем как есть, но логируем
                    logger.warning(f"Could not normalize phone number: {contact.phone}")
            
            # Skip if phone number already exists (duplicate) - проверяем по нормализованному номеру
            if contact.phone:
                normalized = normalize_phone(contact.phone)
                if normalized and normalized in seen_phones:
                    logger.info(f"Skipping duplicate contact with phone: {contact.phone} (normalized: {normalized})")
                    continue
                seen_phones.add(normalized)
            
            existing_contacts.append(contact)
            
        # Save back to file
        await self._save_data(existing_contacts)

    async def update_contact(self, contact_id: int, update_data: Contact):
        """Update an existing contact"""
        contacts = await self.get_contacts()
        
        found = False
        for i, c in enumerate(contacts):
            if c.id == contact_id:
                # Update fields while preserving other attributes
                c.name = update_data.name
                
                # Нормализуем телефон перед обновлением
                if update_data.phone:
                    normalized_phone = normalize_phone(update_data.phone)
                    if normalized_phone:
                        c.phone = normalized_phone
                    else:
                        # Если нормализация не удалась, оставляем как есть, но логируем
                        logger.warning(f"Could not normalize phone number: {update_data.phone}, keeping original")
                        c.phone = update_data.phone
                else:
                    c.phone = update_data.phone
                
                c.group = update_data.group
                
                # Update enrichment fields if present
                if update_data.avatar_url:
                    c.avatar_url = update_data.avatar_url
                if update_data.whatsapp_name:
                    c.whatsapp_name = update_data.whatsapp_name
                if update_data.whatsapp_id:
                    c.whatsapp_id = update_data.whatsapp_id
                
                # Update additional WhatsApp fields
                if update_data.whatsapp_email is not None:
                    c.whatsapp_email = update_data.whatsapp_email
                if update_data.whatsapp_category is not None:
                    c.whatsapp_category = update_data.whatsapp_category
                if update_data.whatsapp_description is not None:
                    c.whatsapp_description = update_data.whatsapp_description
                if update_data.whatsapp_is_business is not None:
                    c.whatsapp_is_business = update_data.whatsapp_is_business
                if update_data.whatsapp_last_seen is not None:
                    c.whatsapp_last_seen = update_data.whatsapp_last_seen
                if update_data.whatsapp_products is not None:
                    c.whatsapp_products = update_data.whatsapp_products
                if update_data.whatsapp_is_registered is not None:
                    c.whatsapp_is_registered = update_data.whatsapp_is_registered
                
                # КРИТИЧЕСКИ ВАЖНО: Обновляем extracted_info ТОЛЬКО если оно передано явно
                # НЕ перезаписываем, если update_data.extracted_info is None (это означает, что поле не должно обновляться)
                # Обновляем только если это не None (включая пустую строку "")
                if update_data.extracted_info is not None:
                    # ДОПОЛНИТЕЛЬНАЯ ПРОВЕРКА: убеждаемся, что мы обновляем правильный контакт
                    if c.id != contact_id:
                        logger.error(f"❌ CRITICAL ERROR: Contact ID mismatch! Expected {contact_id}, but found {c.id}")
                        raise ValueError(f"Contact ID mismatch: expected {contact_id}, got {c.id}")
                    
                    old_length = len(c.extracted_info) if c.extracted_info else 0
                    c.extracted_info = update_data.extracted_info
                    new_length = len(c.extracted_info) if c.extracted_info else 0
                    logger.info(f"✅ Updated extracted_info for contact {contact_id} ({c.name}): {old_length} -> {new_length} chars")
                # Если extracted_info is None, НЕ обновляем - сохраняем существующее значение
                
                # Обновляем draft_info (черновик) ТОЛЬКО если оно передано явно
                if update_data.draft_info is not None:
                    # ДОПОЛНИТЕЛЬНАЯ ПРОВЕРКА: убеждаемся, что мы обновляем правильный контакт
                    if c.id != contact_id:
                        logger.error(f"❌ CRITICAL ERROR: Contact ID mismatch! Expected {contact_id}, but found {c.id}")
                        raise ValueError(f"Contact ID mismatch: expected {contact_id}, got {c.id}")
                    
                    old_length = len(c.draft_info) if c.draft_info else 0
                    c.draft_info = update_data.draft_info
                    new_length = len(c.draft_info) if c.draft_info else 0
                    logger.info(f"✅ Updated draft_info for contact {contact_id} ({c.name}): {old_length} -> {new_length} chars")
                # Если draft_info is None, НЕ обновляем - сохраняем существующее значение
                
                # Preserve last_sync_at if it exists in update_data
                if hasattr(update_data, 'last_sync_at') and update_data.last_sync_at:
                    c.last_sync_at = update_data.last_sync_at
                
                # Обновляем analyzed_message_ids (список проанализированных сообщений)
                if update_data.analyzed_message_ids is not None:
                    c.analyzed_message_ids = update_data.analyzed_message_ids
                
                # ФИНАЛЬНАЯ ПРОВЕРКА: убеждаемся, что ID не изменился
                if c.id != contact_id:
                    logger.error(f"❌ CRITICAL ERROR: Contact ID changed during update! Expected {contact_id}, but got {c.id}")
                    raise ValueError(f"Contact ID changed during update: expected {contact_id}, got {c.id}")
                
                contacts[i] = c
                found = True
                
                # Логирование уже выполнено выше при обновлении полей (строки 169 и 186)
                # Дополнительное логирование не требуется
                
                break
        
        if not found:
            raise HTTPException(status_code=404, detail=f"Contact {contact_id} not found")
        
        # Save updated contacts (удален дубликат вызова)
        await self._save_data(contacts)
        
        # Возвращаем обновленный контакт (используем найденный контакт из цикла)
        updated_contact = next((c for c in contacts if c.id == contact_id), None)
        if updated_contact:
            return updated_contact
        else:
            raise HTTPException(status_code=404, detail=f"Contact {contact_id} not found after update")

    async def delete_contact(self, contact_id: int):
        """Delete a contact by ID"""
        contacts = await self.get_contacts()
        updated_contacts = [c for c in contacts if c.id != contact_id]
        
        if len(updated_contacts) == len(contacts):
            raise HTTPException(status_code=404, detail=f"Contact {contact_id} not found")
        
        await self._save_data(updated_contacts)

    async def bulk_move(self, contact_ids: List[int], target_group: str):
        """Move multiple contacts to a target group"""
        contacts = await self.get_contacts()
        updated_count = 0
        
        for contact in contacts:
            if contact.id in contact_ids:
                contact.group = target_group
                updated_count += 1
                
        if updated_count > 0:
            await self._save_data(contacts)
            
        return {"updated": updated_count}

    async def bulk_delete(self, contact_ids: List[int]):
        """Delete multiple contacts"""
        contacts = await self.get_contacts()
        original_count = len(contacts)
        
        updated_contacts = [c for c in contacts if c.id not in contact_ids]
        deleted_count = original_count - len(updated_contacts)
        
        if deleted_count > 0:
            await self._save_data(updated_contacts)
            
        return {"deleted": deleted_count}
    
    async def bulk_update_whatsapp_ids(self) -> Dict:
        """
        Массовое обновление WhatsApp ID для всех контактов через Green API.
        Возвращает статистику обновлений.
        """
        from app.services.green_api_service import green_api_service
        
        contacts = await self.get_contacts()
        
        updated_count = 0
        failed_count = 0
        skipped_count = 0
        
        for contact in contacts:
            try:
                # Пропускаем контакты, у которых уже есть WhatsApp ID
                if contact.whatsapp_id:
                    skipped_count += 1
                    continue
                
                # Проверяем, зарегистрирован ли номер в WhatsApp
                whatsapp_check = await green_api_service.check_whatsapp(contact.phone)
                contact.whatsapp_is_registered = whatsapp_check.get("exists", False)
                
                # Получаем информацию о контакте через Green API
                contact_info = await green_api_service.get_contact_info(contact.phone)
                
                if contact_info.get("exists"):
                    # Обновляем WhatsApp ID
                    if contact_info.get("whatsapp_id"):
                        contact.whatsapp_id = contact_info["whatsapp_id"]
                    
                    # Получаем возможное имя из WhatsApp
                    whatsapp_name = contact_info.get("name")
                    
                    # Обновляем WhatsApp имя
                    if whatsapp_name and whatsapp_name != contact.phone:
                        if not contact.whatsapp_name or contact.whatsapp_name == contact.phone:
                            contact.whatsapp_name = whatsapp_name
                    
                    # Заменяем основное имя, если оно пустое или равно номеру телефона
                    if whatsapp_name and whatsapp_name != contact.phone:
                        # Нормализуем телефон для сравнения
                        normalized_phone = normalize_phone(contact.phone)
                        normalized_name = normalize_phone(contact.name) if contact.name else ""
                        
                        # Проверяем, является ли имя пустым или номером телефона
                        is_name_empty = not contact.name or not contact.name.strip()
                        is_name_phone = (normalized_name == normalized_phone) or (contact.name == contact.phone)
                        
                        if is_name_empty or is_name_phone:
                            old_name = contact.name or "(пусто)"
                            contact.name = whatsapp_name
                            logger.info(f"✅ Заменено имя контакта {contact.id}: '{old_name}' -> '{whatsapp_name}' (было пустое или номер телефона)")
                    
                    # Обновляем все дополнительные поля
                    if contact_info.get("email"):
                        contact.whatsapp_email = contact_info["email"]
                    if contact_info.get("category"):
                        contact.whatsapp_category = contact_info["category"]
                    if contact_info.get("description"):
                        contact.whatsapp_description = contact_info["description"]
                    if contact_info.get("isBusiness") is not None:
                        contact.whatsapp_is_business = contact_info["isBusiness"]
                    if contact_info.get("lastSeen"):
                        contact.whatsapp_last_seen = contact_info["lastSeen"]
                    if contact_info.get("products"):
                        contact.whatsapp_products = contact_info["products"]
                    
                    updated_count += 1
                    logger.info(f"✅ Updated contact {contact.id} ({contact.name}) with WhatsApp ID: {contact_info.get('whatsapp_id', 'N/A')}")
                else:
                    failed_count += 1
                    logger.debug(f"Could not get WhatsApp info for contact {contact.id} ({contact.name})")
                    
            except Exception as e:
                failed_count += 1
                logger.error(f"Error updating WhatsApp ID for contact {contact.id} ({contact.name}): {e}", exc_info=True)
        
        # Сохраняем обновленные контакты
        if updated_count > 0:
            await self._save_data(contacts)
        
        return {
            "updated": updated_count,
            "failed": failed_count,
            "skipped": skipped_count,
            "total": len(contacts)
        }

    async def _save_data(self, contacts: List[Contact]):
        """Helper to save contacts to JSON file using direct write to avoid PermissionError"""
        data = [c.model_dump() for c in contacts]
        
        # Проверяем, что extracted_info сохраняется
        for contact in contacts:
            if contact.id in [595, 779]:  # Дмитрий и Андрей
                extracted_len = len(contact.extracted_info) if contact.extracted_info else 0
                logger.info(f"💾 Saving contact {contact.id} ({contact.name}): extracted_info length = {extracted_len}")
        
        try:
            # Use direct write instead of temp file + rename to avoid PermissionError on Windows
            async with aiofiles.open(self.file_path, "w", encoding="utf-8") as f:
                await f.write(json.dumps(data, ensure_ascii=False, indent=2, default=str))
            logger.info(f"✅ Successfully saved {len(contacts)} contacts to {self.file_path}")
        except Exception as e:
            logger.error(f"❌ Error saving contacts: {e}", exc_info=True)
            raise e

    async def parse_text_with_ai(self, text: str) -> List[Dict[str, str]]:
        if not self.configured:
            print("LLM is not configured")
            return []

        prompt = f"""
        Extract names and phone numbers from the following text.
        Return a JSON array of objects, where each object has 'name' and 'phone' keys.
        If a phone number is missing, try to find it. If a name is missing, use a placeholder or context.
        Normalize phone numbers to international format if possible.
        
        Text:
        {text}
        
        Output JSON:
        """
        
        try:
            response = await self.model.generate_content_async(
                prompt,
                generation_config=genai.types.GenerationConfig(
                    response_mime_type="application/json"
                )
            )
            
            content = response.text
            result = json.loads(content)
            
            # Handle potential variations in JSON structure (e.g. {"contacts": [...]})
            if "contacts" in result:
                return result["contacts"]
            elif isinstance(result, list):
                return result
            else:
                # Try to find a list in the values
                for key, value in result.items():
                    if isinstance(value, list):
                        return value
                return []
                
        except Exception as e:
            print(f"Error parsing contacts with AI: {e}")
            return []

    async def parse_xlsx_file(self, file_contents: bytes) -> List[Dict[str, str]]:
        """Parse XLSX file and extract contacts"""
        import io
        from openpyxl import load_workbook
        
        try:
            # Load workbook from bytes
            workbook = load_workbook(io.BytesIO(file_contents))
            sheet = workbook.active
            
            contacts = []
            # Assuming first row is headers, skip it
            # Expected columns: Name, Phone (or similar)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):  # Skip empty rows
                    continue
                
                # Try to extract name and phone from first two columns
                name = str(row[0]) if row[0] else ""
                phone = str(row[1]) if len(row) > 1 and row[1] else ""
                
                # Clean name and normalize phone
                name = clean_name(name)
                phone = normalize_phone(phone)
                
                if name or phone:  # At least one field should be present
                    contacts.append({
                        "name": name,
                        "phone": phone
                    })
            
            return contacts
            
        except Exception as e:
            print(f"Error parsing XLSX file: {e}")
            raise HTTPException(status_code=400, detail=f"Error parsing XLSX file: {str(e)}")

contact_service = ContactService()
